"""
TORCH Tools — File Operations
Search, read, move, delete, and manage files.
"""

import os
import shutil
import zipfile
import logging
import threading
import time
from pathlib import Path
from typing import Optional, List, Dict, Any

from config.settings import settings

logger = logging.getLogger("torch.tools.files")

_SKIP_DIRS = {
    'node_modules', '__pycache__', '.git', 'venv', '.venv',
    'Windows', 'Program Files', '$Recycle.Bin', 'AppData', 'System Volume Information',
}
_MAX_INDEX_ENTRIES = 20000

# In-memory index cache: root -> {"built_at": float, "entries": [{path, mtime, size}]}
_FILE_INDEX_CACHE: Dict[str, Dict[str, Any]] = {}
_INDEX_LOCK = threading.Lock()


def _index_ttl() -> float:
    return max(1, int(settings.file_index_ttl_seconds))


def _load_index_for(root: Path) -> List[Dict[str, Any]]:
    """Return cached file metadata for a directory, rebuilding it lazily.

    The first search on a root walks it once (exactly like the old code);
    repeat searches within the TTL reuse the cached list, which keeps
    multi-step tasks and repeated lookups fast.
    """
    root_str = str(root)
    now = time.time()
    with _INDEX_LOCK:
        cached = _FILE_INDEX_CACHE.get(root_str)
        if cached and (now - cached["built_at"]) < _index_ttl():
            return cached["entries"]

    entries: List[Dict[str, Any]] = []
    try:
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames[:] = [d for d in dirnames if not d.startswith('.') and d not in _SKIP_DIRS]
            for f in filenames:
                full_path = os.path.join(dirpath, f)
                try:
                    st = os.stat(full_path)
                except OSError:
                    continue
                entries.append({"path": full_path, "mtime": st.st_mtime, "size": st.st_size})
                if len(entries) >= _MAX_INDEX_ENTRIES:
                    break
            if len(entries) >= _MAX_INDEX_ENTRIES:
                break
    except PermissionError:
        pass

    with _INDEX_LOCK:
        _FILE_INDEX_CACHE[root_str] = {"built_at": now, "entries": entries}
    return entries


def _clear_index_for(root: Optional[Path] = None) -> None:
    """Drop cached entries (used by tests and after destructive file ops)."""
    with _INDEX_LOCK:
        if root is None:
            _FILE_INDEX_CACHE.clear()
        else:
            _FILE_INDEX_CACHE.pop(str(root), None)

def _normalize_find_inputs(name: str, path: str) -> tuple[str, Path]:
    clean_name = str(name or "").strip()
    clean_path = str(path or "").strip()
    if not clean_name or clean_name in {".", ".."}:
        raise ValueError("File name cannot be empty or malformed")
    if "\x00" in clean_name or "/" in clean_name or "\\" in clean_name:
        raise ValueError("File name must not contain a path")
    if not clean_path or "\x00" in clean_path:
        raise ValueError("Search path cannot be empty or malformed")

    expanded = Path(clean_path).expanduser()
    if ".." in expanded.parts:
        raise ValueError("Search path traversal is not allowed")
    normalized = expanded.resolve()
    if not normalized.exists() or not normalized.is_dir():
        raise ValueError(f"Search path is not a valid directory: {normalized}")
    return clean_name, normalized


def find_file(name: str, path: str = "~", limit: int = 20, recent: bool = True) -> str:
    """Recursively search for a file by name, newest modified first.

    ``limit`` caps the number of results and ``recent`` (default True) sorts
    matches by modification time so the most recently edited file is listed
    first — which is what most "find the latest X" requests need.
    """
    name, search_path = _normalize_find_inputs(name, path)
    logger.info(f"Searching for '{name}' in {search_path}")

    matches = []
    for entry in _load_index_for(search_path):
        if name.lower() in os.path.basename(entry["path"]).lower():
            matches.append(entry)

    if recent:
        matches.sort(key=lambda e: e["mtime"], reverse=True)

    max_results = max(1, int(limit or 20))
    matches = matches[:max_results]

    if not matches:
        return f"No files matching '{name}' found in {search_path}"

    result_lines = [f"Found {len(matches)} file(s):"]
    for m in matches:
        size_str = _format_size(m["size"])
        result_lines.append(f"  {m['path']} ({size_str})")

    return "\n".join(result_lines)


def read_pdf(filepath: str) -> str:
    """Extract text from a PDF file."""
    import pdfplumber

    filepath = Path(filepath).expanduser().resolve()
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    text_parts = []
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                text_parts.append(f"--- Page {i + 1} ---\n{text}")

    if not text_parts:
        return "PDF contains no extractable text"

    return "\n\n".join(text_parts)


def read_word(filepath: str) -> str:
    """Extract text from a Word document."""
    from docx import Document

    filepath = Path(filepath).expanduser().resolve()
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    doc = Document(str(filepath))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else "Document is empty"


def read_excel(filepath: str) -> str:
    """Extract data from an Excel file."""
    from openpyxl import load_workbook

    filepath = Path(filepath).expanduser().resolve()
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        results.append(f"Sheet: {sheet_name}")
        for row in rows[:50]:  # Limit rows
            results.append("  " + " | ".join(str(c) if c is not None else "" for c in row))

    wb.close()
    return "\n".join(results) if results else "Spreadsheet is empty"


def move_file(src: str, dst: str) -> str:
    """Move a file from source to destination."""
    src_path = Path(src).expanduser().resolve()
    dst_path = Path(dst).expanduser().resolve()

    if not src_path.exists():
        raise FileNotFoundError(f"Source not found: {src_path}")

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src_path), str(dst_path))
    _clear_index_for(dst_path.parent if dst_path.parent.exists() else None)
    return f"Moved {src_path} → {dst_path}"


def delete_file(filepath: str) -> str:
    """Delete a file. Always requires HITL approval."""
    path = Path(filepath).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if path.is_dir():
        shutil.rmtree(path)
        _clear_index_for(path.parent if path.parent.exists() else None)
        return f"Deleted directory: {path}"
    else:
        path.unlink()
        _clear_index_for(path.parent if path.parent.exists() else None)
        return f"Deleted file: {path}"


def create_folder(path: str) -> str:
    """Create a new directory."""
    dir_path = Path(path).expanduser().resolve()
    if dir_path.exists():
        if dir_path.is_dir():
            return f"Folder already exists: {dir_path}"
        raise RuntimeError(f"A file already exists at: {dir_path}")
    dir_path.mkdir(parents=True, exist_ok=False)
    return f"Created folder: {dir_path}"


def zip_files(files: List[str], output: str) -> str:
    """Compress files into a zip archive."""
    output_path = Path(output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(str(output_path), 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            fp = Path(f).expanduser().resolve()
            if fp.exists():
                zf.write(str(fp), fp.name)

    return f"Created archive: {output_path}"


def _format_size(size_bytes: int) -> str:
    """Format bytes to human-readable size."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f"{size_bytes:.1f}{unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f}TB"


def find_file_fuzzy(name: str, path: str = "~", limit: int = 3) -> dict:
    """
    Search for a file with fuzzy matching when exact match fails.
    Returns: { "found": str|None, "suggestions": list[str], "exact": bool }
    """
    import difflib
    from pathlib import Path

    name, search_path = _normalize_find_inputs(name, path)
    entries = _load_index_for(search_path)

    # Clean input name
    search_name = name.lower()
    search_stem = Path(search_name).stem.lower()

    # 1. Check for exact substring matches first (case-insensitive)
    exact_matches = [
        e for e in entries if search_name in os.path.basename(e["path"]).lower()
    ]

    if exact_matches:
        # Newest, shortest-name matches first — the most "exact" recent file.
        exact_matches.sort(key=lambda e: (len(os.path.basename(e["path"])), -e["mtime"]))
        return {
            "found": exact_matches[0]["path"],
            "suggestions": [e["path"] for e in exact_matches[:limit]],
            "exact": True,
        }

    # 2. Fuzzy match using difflib
    fuzzy_matches = []
    for entry in entries:
        fname = os.path.basename(entry["path"])
        fstem = Path(fname).stem.lower()

        # Compare stems (filenames without extensions)
        ratio = difflib.SequenceMatcher(None, search_stem, fstem).ratio()

        # Also check if search_stem is a substring of fstem or vice versa
        if search_stem in fstem or fstem in search_stem:
            ratio = max(ratio, 0.8)

        if ratio > 0.6:
            fuzzy_matches.append((ratio, entry))

    # Sort by ratio (highest first), then newest modified first
    fuzzy_matches.sort(key=lambda x: (x[0], x[1]["mtime"]), reverse=True)
    top_matches = [e["path"] for _, e in fuzzy_matches[:max(1, int(limit or 3))]]

    return {
        "found": top_matches[0] if top_matches else None,
        "suggestions": top_matches,
        "exact": False,
    }


def list_directory(path: str = "~") -> str:
    """List all files in a directory in a formatted way."""
    dir_path = Path(path).expanduser().resolve()
    if not dir_path.exists():
        return f"Directory not found: {dir_path}"
    if not dir_path.is_dir():
        return f"Not a directory: {dir_path}"

    try:
        files = list(dir_path.iterdir())
    except PermissionError:
        return f"Permission denied accessing {dir_path}"

    if not files:
        return f"Directory {dir_path} is empty."

    # Sort: directories first, then files by name
    dirs = sorted([f for f in files if f.is_dir()], key=lambda x: x.name.lower())
    docs = sorted([f for f in files if f.is_file()], key=lambda x: x.name.lower())

    lines = [f"Contents of {dir_path}:"]
    
    if dirs:
        lines.append("\nDirectories:")
        for d in dirs[:15]:
            lines.append(f"  [DIR]  {d.name}")
        if len(dirs) > 15:
            lines.append(f"  ... and {len(dirs) - 15} more directories")

    if docs:
        lines.append("\nFiles:")
        for f in docs[:30]:
            size_str = _format_size(f.stat().st_size)
            lines.append(f"  {f.name} ({size_str})")
        if len(docs) > 30:
            lines.append(f"  ... and {len(docs) - 30} more files")

    return "\n".join(lines)
